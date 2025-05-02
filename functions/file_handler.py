import pandas as pd
import openpyxl as xl
import openpyxl as xl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import *
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def read_excel_safely(file_path):
    """엑셀 파일 경로로부터 시트 목록을 안전하게 반환"""
    if not file_path.lower().endswith((".xlsx", ".xls")):
        raise ValueError("엑셀 파일이 아닙니다. 확장자 확인 요망.")

    try:
        xl = pd.ExcelFile(file_path)
        return xl.sheet_names
    except Exception as e:
        raise RuntimeError(f"엑셀 파일 읽기 오류: {str(e)}")
    

def set_columns(file_path, sheet_name):
    try:
        target_file = pd.read_excel(file_path, sheet_name = sheet_name)
        print(target_file.columns)
        return list(target_file.columns)
    except Exception as e:
        raise e

def set_spe_columns(file_path, sheet_name):
    print(file_path, sheet_name)
    try:
        target_file = pd.read_excel(file_path, sheet_name = sheet_name, skiprows = 2)
        print(target_file.columns)
        return list(target_file.columns)
    except Exception as e:
        print("여기서 에러")
        raise e

def divide_location(div_dict, label):
    print(div_dict)
    target_name = sorted(['서울', '경기', '인천', '경북', '경남', '전남', '전북', '충남', '충북', '대전세종', '광주', '대구', '울산', '부산', '강원', '제주'])
    
    df = pd.read_excel(div_dict["file_path"], 
                       sheet_name = div_dict["sheet_name"], 
                       dtype = str)
    target_file = div_dict["file_path"]
    address = div_dict["address"]
    share_num = div_dict["share_num"]
    com_name = div_dict["com_name"]
    df[share_num] = df[share_num].astype(int)
    
    for idx, i in enumerate(target_name):
        temp_loc = ""
        if i == "서울":
            temp_loc = "서울특별시"
        elif i == "경기":
            temp_loc = "경기도"
        elif i == "인천":
            temp_loc = "인천광역시"
        elif i == "부산":
            temp_loc = "부산광역시"
        elif i == "경남":
            temp_loc = "경상남도"
        elif i == "경북":
            temp_loc = "경상북도"
        elif i == "대구":
            temp_loc = "대구광역시"
        elif i == "울산":
            temp_loc = "울산광역시"
        elif i == "전남":
            temp_loc = "전라남도"
        elif i == "전북":
            temp_loc = "전라북도"
        elif i == "광주":
            temp_loc = "광주광역시"
        elif i == "충남":
            temp_loc = "충청남도"
        elif i == "충북":
            temp_loc = "충청북도"
        elif i == "대전세종":
            temp_df = df[df[address].str.startswith('대전') | df[address].str.startswith('세종') | 
                        df[address].str.contains("대전광역시") | df[address].str.contains("세종특별자치시")]
            temp_df = temp_df.sort_values(by = share_num, ascending = False).reset_index(drop = True)
            # temp_df = temp_df.drop(['권리기준일자', '주식종류', '우편번호', '법인구분', '내외국인구분'], axis = 1)
            temp_df.index = temp_df.index + 1
            with pd.ExcelWriter(target_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                temp_df.to_excel(writer, sheet_name=i, startrow=2, index=True)  # 3번째 행부터 데이터 저장

                # 🔹 제목 추가 (A1:C1 병합 후 텍스트 삽입)
                workbook = writer.book
                worksheet = writer.sheets[i]

                title_text =    f"회사명 : {com_name}   지역 : {i}"
                title_cell = worksheet.cell(row=1, column=1, value=title_text)

                # 🔹 스타일 적용
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")

                # 🔹 A1:C1 병합
                last_col = get_column_letter(temp_df.shape[1] + 1)  # 마지막 컬럼 계산
                worksheet.merge_cells(f"A1:{last_col}1")

            continue
        
        elif i == "제주":
            temp_loc = "제주특별자치도"
        elif i == "강원":
            temp_loc = "강원도"
        
        
        temp_df = df[df[address].str.startswith(i) | df[address].str.contains(temp_loc)]
        # temp_df = temp_df.drop(['권리기준일자', '주식종류', '우편번호', '법인구분', '내외국인구분'], axis = 1)
        temp_df = temp_df.sort_values(by= share_num, ascending=False).reset_index(drop=True)
        temp_df.index = temp_df.index + 1  # 인덱스 1부터 시작

        # 🔹 openpyxl 엔진을 사용하여 엑셀 저장
        with pd.ExcelWriter(target_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            temp_df.to_excel(writer, sheet_name=i, startrow=2, index=True)  # 3번째 행부터 데이터 저장

            # 🔹 제목 추가 (A1:C1 병합 후 텍스트 삽입)
            workbook = writer.book
            worksheet = writer.sheets[i]

            title_text = f"회사명 : {com_name}   지역 : {i}"
            title_cell = worksheet.cell(row=1, column=1, value=title_text)

            # 🔹 스타일 적용
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # 🔹 A1:C1 병합
            last_col = get_column_letter(temp_df.shape[1] + 1)  # 마지막 컬럼 계산
            worksheet.merge_cells(f"A1:{last_col}1")
            
            percent = float((idx + 1) / len(target_name)) * 100
            label.setText(f"현재 {i} 지역 처리완료.. {percent} % 완료 되었습니다.")
            QApplication.processEvents()
            print(i, " 지역 처리중")


def divide_spe_location(div_dict, label):
    file_path = div_dict["file_path"]
    sheet_name = div_dict["sheet_name"]
    target_dict = div_dict["target_dict"]
    com_name = div_dict["com_name"]
    share_num = div_dict["share_num"]
    address = div_dict["address"]
    target_sum = sum(len(v) for v in target_dict.values())
    count = 0
    target_df = pd.read_excel(file_path, sheet_name = sheet_name, dtype = str, skiprows = 2)

    target_df[share_num] = target_df[share_num].astype(int)
    target_df = target_df.drop(columns = "Unnamed: 0")
    for k, v_list in target_dict.items():
        temp_df = pd.DataFrame(columns = target_df.columns)
        temp_name = f"{k}-{''.join(v_list)}"
        if len(temp_name) > 30:
            temp_sheet_name = k
        else:
            temp_sheet_name = temp_name

        for spe in v_list:
            count += 1
            if "경기" in k and spe == "양주시":
                temp_df = pd.concat([temp_df, target_df[target_df[address].str.contains(rf'(?<!남){spe}')]])
            elif "부산" in k and spe == "서구":
                temp_df = pd.concat([temp_df, target_df[target_df[address].str.contains(rf'(?<!강){spe}')]])
            else:
                temp_df = pd.concat([temp_df, target_df[target_df[address].str.contains(rf'{spe}')]])
            percent = float((count) / target_sum) * 100
            label.setText(f"현재 {k} 지역 {spe} 분류 진행중.. {percent}% 완료")
            QApplication.processEvents()
        temp_df = temp_df.sort_values(by = share_num, ascending = False).reset_index(drop = True)
        temp_df.index = temp_df.index + 1

        with pd.ExcelWriter(file_path, engine = "openpyxl", mode = 'a', if_sheet_exists = "replace") as writer:
            temp_df.to_excel(writer, sheet_name = temp_sheet_name, startrow = 2, index = True)

            workbook = writer.book
            worksheet = writer.sheets[temp_sheet_name]

            title_text = f"회사명 : {com_name}   지역 : {temp_name}"
            title_cell = worksheet.cell(row = 1, column = 1, value = title_text)

            title_cell.font = Font(bold = True, size = 14)
            title_cell.alignment = Alignment(horizontal = "center", vertical = "center")

            last_col = get_column_letter(temp_df.shape[1] + 1)
            worksheet.merge_cells(f"A1:{last_col}1")

